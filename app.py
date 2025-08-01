# app.py – Monthly Costs Processor
# --------------------------------------------------------------
import streamlit as st
import pandas as pd
import io
from openpyxl.utils import get_column_letter
st.set_page_config(page_title="Monthly Costs Processor", layout="wide")
st.title("Monthly Costs Processor")

# ── Session state ──────────────────────────────────────────────
if 'storage_processed' not in st.session_state:
    st.session_state.storage_processed = False
if 'monthly_df' not in st.session_state:
    st.session_state.monthly_df = None
if 'orders_df' not in st.session_state:
    st.session_state.orders_df = None
if 'valid_skus' not in st.session_state:
    st.session_state.valid_skus = None
if 'owner_map' not in st.session_state:
    st.session_state.owner_map = {}

# ── Step 1: Orders Processing ─────────────────────────────────
st.header("Step 1: Orders Processing")
col1, col2 = st.columns(2)

with col1:
    month_year = st.text_input(
        "Month and Year (e.g., Mar-24)",
        key="month_year",
    )
with col2:
    monthly_file = st.file_uploader(
        "Upload Order Line Level Data",
        type=['xlsx'],
    )
if monthly_file:
    try:
        xl = pd.ExcelFile(monthly_file)
        monthly_df = pd.read_excel(monthly_file, sheet_name=xl.sheet_names[0], header=1)

        monthly_df.columns = monthly_df.columns.str.strip()

        if 'Date Ordered' in monthly_df.columns:
            monthly_df['Date Ordered'] = (
                pd.to_datetime(monthly_df['Date Ordered'])
                  .dt.strftime('%Y-%m-%d %H:%M:%S')
            )

        if 'Location Code' in monthly_df.columns:
            monthly_df['Location Code'] = monthly_df['Location Code'].astype(str)

        monthly_df['Stock Code'] = (
            monthly_df['Stock Code'].astype(str).str.strip().str.upper()
        )
        monthly_df['Total Locations'] = (
            pd.to_numeric(monthly_df['Total Locations'], errors='coerce')
              .fillna(0)
        )

        st.session_state.monthly_df = monthly_df
        st.success("Orders file loaded. Continue to Step 2 to process.")

    except Exception as e:
        st.error(f"An error occurred in Orders processing: {e}")
        st.error("Please check that the files have the expected format and try again.")

# ── Step 2: Storage Processing ────────────────────────────────
if st.session_state.monthly_df is not None:
    st.header("Step 2: Storage Processing")
    col1, col2 = st.columns(2)

    with col1:
        stock_file   = st.file_uploader("Upload Stock Report", type=['xlsx'])
    with col2:
        storage_file = st.file_uploader("Upload Staci Storage Tab", type=['xlsx'])

    if stock_file and storage_file:
        try:
            stock_df   = pd.read_excel(stock_file,   header=1)
            storage_df = pd.read_excel(storage_file, header=1)

            stock_df['Stock Code'] = (
                stock_df['Stock Code'].astype(str).str.strip().str.upper()
            )
            stock_df['Event'] = stock_df['Event'].astype(str).str.strip()
            valid_skus = set(
                stock_df.loc[stock_df['Event'] == 'Dummy Cards & Boxes', 'Stock Code']
            )
            st.session_state.valid_skus = valid_skus

            filtered_stock = stock_df[stock_df['Stock Code'].isin(valid_skus)].copy()

            filtered_stock['Pallets'] = filtered_stock['Stock Code'].map(
                storage_df.set_index('Part Number')['Period']
                         .to_dict()
            ).fillna(0)

            filtered_stock['Cost'] = (filtered_stock['Pallets'] * 1.92).round(2)

            monthly_df = st.session_state.monthly_df
            orders_df = monthly_df[monthly_df['Stock Code'].isin(valid_skus)].copy()

            columns_to_remove = [
                'Sell Price', 'Line Status', 'Back Order Status',
                'Back Order Placed Date', 'Sell Price (Packs)'
            ]
            orders_df = orders_df.drop(
                columns=[c for c in columns_to_remove if c in orders_df.columns]
            )

            # Charges
            orders_df['Pick Charge'] = orders_df['Total Locations'] * 1.59

            orders_df['Order_Count'] = (
                orders_df.groupby('Order Number').cumcount() + 1
            )
            orders_df['Packaging'] = (
                (orders_df['Order_Count'] % 10 == 1).astype(int)
                  * orders_df['Total Locations']
            )
            orders_df = orders_df.drop('Order_Count', axis=1)

            orders_df['Packaging Charge'] = orders_df['Packaging'] * 0.97
            orders_df['Label Charge'] = orders_df['Packaging'] * 0.39

            for c in ['Pick Charge', 'Packaging Charge', 'Label Charge']:
                orders_df[c] = orders_df[c].round(2)

            # Map owners and add to orders
            owner_map = filtered_stock.set_index('Stock Code')['Responsible Owner'].to_dict()
            st.session_state.owner_map = owner_map
            desc_col = next(
                (c for c in filtered_stock.columns
                 if c.strip().lower() == 'full description'),
                None
            )
            desc_map = (
                filtered_stock.set_index('Stock Code')[desc_col].to_dict()
                if desc_col
                else {}
            )
            orders_df['Responsible Owner'] = (
                orders_df['Stock Code'].map(owner_map).fillna('Unknown')
            )
            if desc_col:
                insert_idx = (
                    orders_df.columns.get_loc('Stock Title') + 1
                    if 'Stock Title' in orders_df.columns
                    else len(orders_df.columns)
                )
                orders_df.insert(
                    insert_idx,
                    'Full Description',
                    orders_df['Stock Code'].map(desc_map)
                )

            st.header("Orders Processing Results")
            c1, c2, c3 = st.columns(3)
            c1.metric("Total Orders", len(monthly_df))
            c2.metric("Valid Orders", len(orders_df))
            c3.metric("Removed Orders", len(monthly_df) - len(orders_df))

            st.session_state.orders_df = orders_df
            st.session_state.storage_processed = True

        except Exception as e:
            st.error(f"An error occurred in Storage processing: {e}")
            st.error("Please check that all files have the expected format and try again.")
            st.session_state.storage_processed = False
else:
    st.info("Please complete Step 1 (Orders Processing) before proceeding to Step 2.")

# ── Step 3: Goods-In Processing & Report Download ─────────────
if st.session_state.valid_skus is not None and st.session_state.storage_processed:
    st.header("Step 3: Goods In Processing")
    goods_in_file = st.file_uploader("Upload Goods In File", type=['xlsx'])

    if goods_in_file:
        try:
            goods_in_df = pd.read_excel(goods_in_file, header=0)

            # Identify part-number column
            part_no_col = next(
                (c for c in goods_in_df.columns
                 if str(c).lower().replace(' ', '') == 'partno'),
                None
            )
            if part_no_col is None:
                st.error("Could not find the part-number column in the Goods In file.")
                st.stop()

            goods_in_df[part_no_col] = (
                goods_in_df[part_no_col].astype(str).str.strip().str.upper()
            )
            filtered_goods_in = goods_in_df[
                goods_in_df[part_no_col].isin(st.session_state.valid_skus)
            ].copy()

            # Keep expected columns even if no rows match
            filtered_goods_in = filtered_goods_in[
                [part_no_col, 'Part Description', 'Qty', 'No Of Containers']
            ]
            filtered_goods_in.columns = [
                'Stock Code', 'Part Description', 'Qty', 'No of Containers'
            ]
            filtered_goods_in['Cost'] = (
                filtered_goods_in['No of Containers'] * 5.26
            ).round(2)
            filtered_goods_in['Responsible Owner'] = (
                filtered_goods_in['Stock Code']
                    .map(st.session_state.owner_map)
                    .fillna('Unknown')
            )

            # ── Build Excel workbooks per owner ─────────────────
            owners = sorted(
                filtered_stock['Responsible Owner'].fillna('Unknown').unique()
            )
            currency_fmt = '_-£* #,##0.00_-;-£* #,##0.00_-;_-£* "-"??_-;_-@_-'
            outputs = []

            for owner in owners:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    owner_orders = orders_df[orders_df['Responsible Owner'] == owner]
                    owner_orders.to_excel(writer, sheet_name='Orders', index=False)
                    ws_orders = writer.sheets['Orders']

                    total_row = len(owner_orders) + 2
                    for col in ['Pick Charge', 'Packaging Charge', 'Label Charge']:
                        col_letter = get_column_letter(owner_orders.columns.get_loc(col) + 1)
                        ws_orders[f"{col_letter}{total_row}"] = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
                    ws_orders[f"A{total_row}"] = 'Total'

                    for col in ['Pick Charge', 'Packaging Charge', 'Label Charge']:
                        col_letter = get_column_letter(owner_orders.columns.get_loc(col) + 1)
                        for r in range(2, total_row + 1):
                            ws_orders[f"{col_letter}{r}"].number_format = currency_fmt

                    owner_storage = filtered_stock[filtered_stock['Responsible Owner'] == owner]
                    owner_storage.to_excel(writer, sheet_name='Storage', index=False)
                    ws_storage = writer.sheets['Storage']

                    st_total_row = len(owner_storage) + 2
                    st_cost_col = get_column_letter(owner_storage.columns.get_loc('Cost') + 1)
                    ws_storage[f"{st_cost_col}{st_total_row}"] = f"=SUM({st_cost_col}2:{st_cost_col}{st_total_row-1})"
                    ws_storage[f"A{st_total_row}"] = 'Total'

                    for r in range(2, st_total_row + 1):
                        ws_storage[f"{st_cost_col}{r}"].number_format = currency_fmt

                    owner_goods_in = filtered_goods_in[filtered_goods_in['Responsible Owner'] == owner]
                    if not owner_goods_in.empty:
                        owner_goods_in.to_excel(writer, sheet_name='Goods In', index=False)
                        ws_gi = writer.sheets['Goods In']

                        gi_total_row = len(owner_goods_in) + 2
                        gi_cost_col = get_column_letter(owner_goods_in.columns.get_loc('Cost') + 1)
                        ws_gi[f"{gi_cost_col}{gi_total_row}"] = f"=SUM({gi_cost_col}2:{gi_cost_col}{gi_total_row-1})"
                        ws_gi[f"A{gi_total_row}"] = 'Total'

                        for r in range(2, gi_total_row + 1):
                            ws_gi[f"{gi_cost_col}{r}"].number_format = currency_fmt

                    summary_rows = [
                        {
                            'Month': month_year,
                            'Tab Name': 'Orders',
                            'Total Cost': owner_orders[['Pick Charge', 'Packaging Charge', 'Label Charge']].sum().sum(),
                        },
                        {
                            'Month': month_year,
                            'Tab Name': 'Storage',
                            'Total Cost': owner_storage['Cost'].sum(),
                        },
                    ]
                    if not owner_goods_in.empty:
                        summary_rows.append(
                            {
                                'Month': month_year,
                                'Tab Name': 'Goods In',
                                'Total Cost': owner_goods_in['Cost'].sum(),
                            }
                        )
                    summary_df = pd.DataFrame(summary_rows)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    ws_summary = writer.sheets['Summary']
                    for r in range(2, len(summary_df) + 2):
                        ws_summary[f"C{r}"].number_format = currency_fmt

                    for ws in writer.sheets.values():
                        for col_cells in ws.columns:
                            max_len = max(
                                len(str(cell.value)) if cell.value else 0
                                for cell in col_cells
                            )
                            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

                output.seek(0)
                outputs.append((owner, output))

            for owner, data_out in outputs:
                st.download_button(
                    label=f"Download Report for {owner}",
                    data=data_out,
                    file_name=f"Dummy_Products_{month_year}_{owner}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # Info if Goods In tab absent
            if filtered_goods_in.empty:
                st.info(
                    "No matching SKUs found in the Goods In file, "
                    "so that tab was left out of the workbook."
                )

        except Exception as e:
            st.error(f"An error occurred in Goods In processing: {e}")
            st.error("Please check that the file has the expected format and try again.")
else:
    if st.session_state.valid_skus is not None:
        st.info("Please complete Step 2 (Storage Processing) before proceeding to Step 3.")
